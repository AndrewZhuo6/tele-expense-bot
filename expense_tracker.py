from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from telegram import Update
from telegram.ext import Application, CommandHandler, CallbackContext, MessageHandler, filters
import sqlite3
import os

db_name = "expenses.db"
excel = "expense_tracker.xlsx"
token = "8476726893:AAF37w1v63y4nmD2H8R8zAu3BiiIiOwWZyI"

def main():
    create_db()
    app = Application.builder().token(token).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("download", download))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, log))
    app.run_polling()

def create_db():
    db = sqlite3.connect(db_name)
    c = db.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS expenses
                (id INTEGER PRIMARY KEY AUTOINCREMENT,
                date DATETIME DEFAULT CURRENT_TIMESTAMP,
                category TEXT NOT NULL,
                amount REAL NOT NULL,
                description TEXT NOT NULL)''')
    db.commit()
    db.close()

async def start(update: Update, context):
    await update.message.reply_text(
        "Hello! send me your expense in the format:\n"
        "Entry-Type(Expense/Income)\nCategory\nAmount\nDescription (Optional)\n"
        "Or use /download to get your expense file.")

async def log(update: Update, context):
    text = update.message.text.split("\n")
    if len(text) >= 3:
        entry_type = text[0].strip().lower()
        category = text[1].strip()
        try:
            amount = float(text[2].strip())
        except ValueError:
            await update.message.reply_text("Amount must be a number")
            return
        description = " ".join(text[3:]) if len(text) > 3 else category
        if entry_type == "expense":
            amount = -amount
        db = sqlite3.connect(db_name)
        c = db.cursor()
        c.execute("INSERT INTO expenses (category, amount, description) VALUES (?, ?, ?)", (category, amount, description))
        db.commit()
        db.close()
        await update.message.reply_text("Logged to database.")
    else:
        await update.message.reply_text("Invalid format. Please use:\nEntry-Type\nCategory\nAmount\nDescription (Optional)")

async def download(update: Update, context):
    db = sqlite3.connect(db_name)
    c = db.cursor()
    c.execute("SELECT date, category, amount, description FROM expenses")
    rows = c.fetchall()
    db.close()
    if not rows:
        await update.message.reply_text("No expense record found.")
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Expenses"
    headers = ["Date", "Category", "Amount", "Description"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in rows:
        ws.append(row)
    wb.save(excel)
    await update.message.reply_document(document=open(excel, "rb"))

if __name__ == "__main__":
    main()
